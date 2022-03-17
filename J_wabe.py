import requests
import re
from selenium import webdriver
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import Font
import openpyxl
import locale

browser = webdriver.Chrome()
jwave_url = 'https://www.j-wave.co.jp/original/tokiohot100/chart/main.htm'
youtube_url = 'https://www.youtube.com/?gl=JP&hl=ja'
req = requests.get(jwave_url).content
soup = BeautifulSoup(req,'html.parser')

ranking_tag = soup.find_all('div',class_= re.compile(r'song_rank'))
artsit_tag = soup.find_all(('div'),class_=re.compile(r'song_artist'))
music_tag = soup.find_all(('div'),class_=re.compile(r'song_title'))

ranking = [num.text for num in ranking_tag]
artsit = [num.text for num in artsit_tag]
music = [num.text for num in music_tag]

df_Jwave = pd.DataFrame({'Artist':artsit,'Music':music},index=ranking)

url_list = []
for art,mus in zip(df_Jwave['Artist'],df_Jwave['Music']):
    browser.get(youtube_url)
    elem = browser.find_element_by_id('search')
    elem.send_keys(art+' '+mus)
    botton = elem.find_element_by_id('search-icon-legacy')
    botton.click()
    url_list.append(browser.current_url)

df_Jwave = pd.DataFrame({'Artist':artsit,'Music':music,'URL':url_list},index=ranking)
file_name = ('J_wave.xlsx')
df_Jwave.to_excel(file_name)

column_list = ['A1','B1','C1','D1']
wb = openpyxl.load_workbook('J_wave.xlsx')
ws = wb.worksheets[0]

range_num = ws['A1':'C101']
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 40
font = Font(name='メイリオ',size=10)
for p in range_num:
    for col in p:
        ws[col.coordinate].font = font
wb.save(file_name)
browser.close()
wb.close()